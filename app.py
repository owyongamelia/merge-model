import os
from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def copy_sheet(source_sheet, target_sheet):
    # (Same implementation as provided in your code)
    # ... [Copy merged cells, row/column dimensions, styles, etc] ...

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Save uploaded file
        file_a = request.files['file_a']
        upload_path = os.path.join(UPLOAD_FOLDER, file_a.filename)
        file_a.save(upload_path)
        
        # Load workbooks
        wb_a = load_workbook(upload_path, data_only=False)
        wb_b = load_workbook('B.xlsx', data_only=False)  # Server file
        
        # Create merged workbook
        new_wb = Workbook()
        new_wb.remove(new_wb.active)
        
        # Copy sheets
        for name, wb in [('A', wb_a), ('B', wb_b)]:
            ws = wb[name]
            new_ws = new_wb.create_sheet(ws.title)
            copy_sheet(ws, new_ws)
        
        # Save and return result
        output_path = os.path.join(UPLOAD_FOLDER, 'C.xlsx')
        new_wb.save(output_path)
        return send_file(output_path, as_attachment=True)
    
    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
